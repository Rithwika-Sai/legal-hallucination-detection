import os
import json
import datetime
import torch
from transformers import AutoTokenizer, AutoModelForCausalLM, LogitsProcessor, LogitsProcessorList
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── CONFIGURATION ───────────────────────────────────────────────────────────
MODEL_ID    = "mistralai/Mistral-7B-Instruct-v0.2"
JSON_PATH   = "data/test.json"
OUTPUT_XLSX = "output/mistral_results.xlsx"
DEVICE      = "cuda" if torch.cuda.is_available() else "cpu"

# ─── DYNAMIC TEMPERATURE LOGITS PROCESSOR ────────────────────────────────────
class DynamicTemperatureLogitsProcessor(LogitsProcessor):
    """
    Adjusts temperature based on structural position in the generated output.

      Start  (0–25%):  temp=0.7  → fluent, natural phrasing
      Middle (25–75%): temp=0.4  → focused / fact-locked
      End   (75–100%): temp=0.7  → clean conclusion

    Floor clamped at 0.4 (not 0.1) to prevent logit overflow → RLHF refusal.
    """
    def __init__(self, start_token_idx: int, max_new_tokens: int):
        self.start_idx      = start_token_idx
        self.max_new_tokens = max_new_tokens

    def __call__(self, input_ids: torch.LongTensor, scores: torch.FloatTensor) -> torch.FloatTensor:
        generated_so_far = input_ids.shape[1] - self.start_idx
        progress         = generated_so_far / self.max_new_tokens
        current_temp     = 0.4 if 0.25 <= progress <= 0.75 else 0.7
        return scores / current_temp


# ─── DATA LOADER ─────────────────────────────────────────────────────────────
def load_case_data(json_path: str) -> list:
    if not os.path.exists(json_path):
        os.makedirs(os.path.dirname(json_path), exist_ok=True)
        sample_data = [{
            "arguments_and_ratio": (
                "The appellant argued that the sole eye-witness was a deeply interested "
                "party closely related to the deceased. The court verified the case record "
                "and held that relationship is not a ground to discard a witness if their "
                "statement is intrinsically credible and corroborated by medical forensics."
            )
        }]
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(sample_data, f, indent=4)
        print(f"[DATA] Sample JSON created at: {json_path}")

    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


# ─── EXCEL SAVE (append-safe) ─────────────────────────────────────────────────
def save_to_excel(context: str, query: str, result: str, output_path: str):
    """
    Saves inference result to Excel.
    - Creates the file + header row on first run.
    - Appends a new data row on every subsequent run.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # ── Styles ───────────────────────────────────────────────────────────────
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="2F4F8F")
    label_font   = Font(name="Arial", bold=True, size=10)
    label_fill   = PatternFill("solid", start_color="D9E1F2")
    content_font = Font(name="Arial", size=10)
    wrap_align   = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="center")
    thin         = Side(style="thin")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    HEADERS = ["Timestamp", "Model", "Query", "Context", "Result"]
    COL_WIDTHS = [22, 38, 40, 60, 60]

    # ── Load existing or create new workbook ─────────────────────────────────
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        ws = wb.active
        next_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inference Results"

        # Write header row
        for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
            cell            = ws.cell(row=1, column=col, value=h)
            cell.font       = header_font
            cell.fill       = header_fill
            cell.alignment  = center_align
            cell.border     = border
            ws.column_dimensions[cell.column_letter].width = w

        next_row = 2

    # ── Write data row ────────────────────────────────────────────────────────
    row_data = [
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        MODEL_ID,
        query,
        context,
        result,
    ]
    for col, value in enumerate(row_data, start=1):
        cell           = ws.cell(row=next_row, column=col, value=value)
        cell.font      = label_font if col <= 2 else content_font
        cell.fill      = label_fill if col <= 2 else PatternFill()
        cell.alignment = wrap_align
        cell.border    = border

    ws.row_dimensions[next_row].height = 120

    wb.save(output_path)
    print(f"[SAVE] Row {next_row} written → {output_path}")


# ─── MAIN PIPELINE ───────────────────────────────────────────────────────────
if __name__ == "__main__":

    # 1. Load data
    case_data = load_case_data(JSON_PATH)[0]
    context   = case_data.get("arguments_and_ratio", "")
    query     = "What did the court decide about the related eyewitness?"

    SYSTEM_GUARDRAIL = (
        "You are a precision legal analytics tool. "
        "Answer strictly from the provided context. "
        "Do not add any information not present in the context."
    )

    final_prompt = (
        f"<s>[INST] {SYSTEM_GUARDRAIL}\n\n"
        f"Context:\n{context}\n\n"
        f"Question: {query} [/INST]"
    )

    # 2. Load model
    print(f"[LOAD] Initializing Mistral on {DEVICE}...")
    tokenizer = AutoTokenizer.from_pretrained(MODEL_ID)
    tokenizer.pad_token = tokenizer.eos_token

    model = AutoModelForCausalLM.from_pretrained(
        MODEL_ID,
        dtype=torch.bfloat16 if DEVICE == "cuda" else torch.float32,
        device_map="auto" if DEVICE == "cuda" else None,
        low_cpu_mem_usage=True,
        attn_implementation="sdpa" if DEVICE == "cuda" else "eager",
    )
    model.eval()

    # 3. Tokenize
    inputs             = tokenizer(final_prompt, return_tensors="pt").to(model.device)
    start_token_length = inputs["input_ids"].shape[1]
    max_output_tokens  = 180

    dynamic_temp_processor = DynamicTemperatureLogitsProcessor(
        start_token_idx=start_token_length,
        max_new_tokens=max_output_tokens,
    )

    # 4. Generate
    print("\n[INFERENCE] Generating...")
    with torch.no_grad():
        output_ids = model.generate(
            **inputs,
            max_new_tokens=max_output_tokens,
            do_sample=False,
            logits_processor=LogitsProcessorList([dynamic_temp_processor]),
            use_cache=True,
            pad_token_id=tokenizer.eos_token_id,
            repetition_penalty=1.15,
        )

    generated_text = tokenizer.decode(
        output_ids[0][start_token_length:], skip_special_tokens=True
    ).strip()

    # 5. Print
    print("\n" + "═" * 60)
    print("MISTRAL DYNAMIC INFERENCE OUTPUT")
    print("═" * 60)
    print(f"Result:\n{generated_text}")
    print("═" * 60)

    # 6. Save to Excel
    save_to_excel(
        context=context,
        query=query,
        result=generated_text,
        output_path=OUTPUT_XLSX,
    )
